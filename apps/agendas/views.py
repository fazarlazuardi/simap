from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from archives.models import Archive
from dispositions.models import Disposition
from users.models import User, Employee
from services.integrations.gateway_service import WhatsAppService
from services.audit_logs.audit_service import AuditService
from users.decorators import pimpinan_required, staff_or_kabid_or_pimpinan_required
from .models import Agenda


class AgendaForm(forms.ModelForm):
    class Meta:
        model = Agenda
        fields = ['title', 'location', 'description', 'scheduled_at', 'archive', 'attachment']


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def build_agenda_export_workbook(agendas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rekap Agenda BAZNAS'

    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "BADAN AMIL ZAKAT NASIONAL (BAZNAS) KABUPATEN TANGERANG"
    title_cell.font = Font(name='Arial', size=13, bold=True, color='046C4E')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    sub_cell.value = "LAPORAN REKAPITULASI AGENDA KERJA, PENUGASAN & SPPD"
    sub_cell.font = Font(name='Arial', size=11, bold=True, color='1E293B')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:H3')
    meta_cell = ws['A3']
    meta_cell.value = f"Tanggal Unduh: {timezone.now().strftime('%d %B %Y %H:%M')} WIB  |  Total Agenda: {len(agendas)} Record"
    meta_cell.font = Font(name='Arial', size=9, italic=True, color='64748B')
    meta_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.append([])

    headers = [
        'No', 'Waktu & Tanggal', 'Nama Agenda / Kegiatan', 'Lokasi Tujuan',
        'Referensi Arsip / Dokumen', 'Status Agenda', 'Pegawai Ditugaskan', 'Notulensi / Hasil'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='046C4E', end_color='046C4E', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[5].height = 26

    even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    for idx, agenda in enumerate(agendas, 1):
        row_num = 5 + idx
        waktu = agenda.scheduled_at.strftime('%d/%m/%Y %H:%M WIB') if agenda.scheduled_at else '-'
        lokasi = agenda.location or '-'
        
        archive_ref = f"{agenda.archive.archive_number} - {agenda.archive.title}" if (agenda.archive and agenda.archive.archive_number) else (agenda.archive.title if agenda.archive else 'Agenda Internal')
        status_str = agenda.get_status_display().upper() if hasattr(agenda, 'get_status_display') else agenda.status.upper()
        
        assigned_names = []
        if agenda.assigned_employees.exists():
            assigned_names = [e.full_name for e in agenda.assigned_employees.all()]
        elif agenda.assigned_to.exists():
            assigned_names = [u.get_full_name() or u.username for u in agenda.assigned_to.all()]
        
        pegawai_str = ', '.join(assigned_names) if assigned_names else '-'
        notulensi = agenda.completed_notes[:100] if agenda.completed_notes else ('Berkas Terunggah' if agenda.completed_file else '-')

        ws.append([
            idx,
            waktu,
            agenda.title,
            lokasi,
            archive_ref,
            status_str,
            pegawai_str,
            notulensi
        ])

        fill = even_fill if idx % 2 == 0 else odd_fill
        ws.row_dimensions[row_num].height = 22

        for col_idx in range(1, 9):
            c = ws.cell(row=row_num, column=col_idx)
            c.fill = fill
            c.border = thin_border
            c.font = Font(name='Arial', size=9)
            if col_idx in [1, 2, 6]:
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    col_widths = {1: 6, 2: 22, 3: 35, 4: 25, 5: 30, 6: 18, 7: 30, 8: 35}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb


@login_required
def agenda_list(request):
    query = request.GET.get('q')
    status = request.GET.get('status')
    disposer = request.GET.get('disposer')
    terkait_user = request.GET.get('terkait_user')
    archive_type = request.GET.get('type')

    # Update otomatis status agenda yang arsipnya sudah selesai
    Agenda.objects.filter(archive__status='selesai', status='terjadwal').update(is_completed=True, status='selesai')

    # Backfill SPPD lama ke agenda
    try:
        from sppd_service.models import SPPD
        from datetime import datetime, time
        existing_sppds = SPPD.objects.select_related('disposition__archive', 'created_by').prefetch_related('assigned_employees').filter(is_cancelled=False)
        for sppd_obj in existing_sppds:
            num_str = sppd_obj.sppd_number
            arc = sppd_obj.disposition.archive if (sppd_obj.disposition and sppd_obj.disposition.archive) else None
            agenda_target = Agenda.objects.filter(description__icontains=num_str).first()
            
            if not agenda_target:
                sch_date = sppd_obj.departure_date
                if sch_date:
                    raw_dt = datetime.combine(sch_date, time(8, 0))
                    try:
                        sch_dt = timezone.make_aware(raw_dt)
                    except Exception:
                        sch_dt = raw_dt
                    
                    purp = sppd_obj.purpose or sppd_obj.destination or "Perjalanan Dinas"
                    agenda_target = Agenda.objects.create(
                        title=f"SPPD: {purp[:50]}",
                        location=sppd_obj.destination,
                        description=f"Perjalanan Dinas SPPD {num_str} ke {sppd_obj.destination}. Maksud: {purp}",
                        archive=arc,
                        scheduled_at=sch_dt,
                        created_by=sppd_obj.created_by or request.user,
                        status='terjadwal',
                    )
            
            if agenda_target:
                if sppd_obj.assigned_employees.exists():
                    agenda_target.assigned_employees.set(sppd_obj.assigned_employees.all())
                elif sppd_obj.surat_tugas and sppd_obj.surat_tugas.pegawai_ditugaskan.exists():
                    agenda_target.assigned_employees.set(sppd_obj.surat_tugas.pegawai_ditugaskan.all())
    except Exception as e:
        pass


def is_superadmin_or_kabid_4(user):
    """
    Hanya Superadmin dan Kabid IV (serta POV Kabid 4 / Superadmin) yang diizinkan 
    melakukan penambahan, pengeditan, atau pengelolaan agenda di Modul Agenda.
    Pengguna lain hanya memiliki akses Read Only.
    """
    if not user or not user.is_authenticated:
        return False
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_superadmin', False):
        return True
    
    # Check active POV in session / attribute
    active_pov = getattr(user, 'active_pov', '')
    if active_pov in ['superadmin', 'kabid_4']:
        return True

    if getattr(user, 'is_kabid_4', False):
        return True

    is_kabid = getattr(user, 'is_kabid', False)
    bidang = getattr(user, 'active_bidang', '')
    if not bidang and hasattr(user, 'employee') and user.employee:
        bidang = getattr(user.employee, 'bidang', '')

    if is_kabid and bidang == 'bidang_4':
        return True

    return False


@login_required
def agenda_list(request):
        'assigned_to',
        Prefetch(
            'archive__dispositions',
            queryset=Disposition.objects.prefetch_related('forwarded_to')
        )
    ).all().order_by('-scheduled_at')

    if query:
        agendas = agendas.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )
    
    if status in ['completed', 'selesai']:
        agendas = agendas.filter(status='selesai')
    elif status in ['upcoming', 'terjadwal']:
        agendas = agendas.filter(status='terjadwal')
    elif status in ['cancelled', 'dibatalkan']:
        agendas = agendas.filter(status='dibatalkan')
    elif status in ['postponed', 'diundur']:
        agendas = agendas.filter(status='diundur')

    if disposer:
        agendas = agendas.filter(
            archive__dispositions__sender_id=disposer
        ).distinct()

    if archive_type:
        agendas = agendas.filter(archive__archive_type=archive_type)

    if terkait_user:
        agendas = agendas.filter(
            Q(archive__dispositions__forwarded_to__user_account__id=terkait_user) |
            Q(archive__dispositions__sppd_list__assigned_employees__user_account__id=terkait_user) |
            Q(assigned_to__id=terkait_user)
        ).distinct()

    disposer_users = User.objects.filter(
        sent_dispositions__isnull=False
    ).distinct().order_by('username')

    terkait_users = User.objects.filter(
        Q(employee__received_dispositions__isnull=False) |
        Q(employee__sppd_assignments__isnull=False)
    ).distinct().order_by('username')

    if 'export' in request.GET:
        wb = build_agenda_export_workbook(agendas)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Agenda_BAZNAS_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    per_page = int(request.GET.get('per_page', 25))
    paginator = Paginator(agendas, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prefetch InternalMeeting for agendas on current page & build safe notulensi JSON
    import json
    meeting_ids = [a.internal_meeting_id for a in page_obj if a.internal_meeting_id]
    meetings_map = {}
    if meeting_ids:
        from internal_meetings.models import InternalMeeting
        meetings = InternalMeeting.objects.filter(pk__in=meeting_ids).select_related(
            'notulis', 'leader'
        ).prefetch_related('participants', 'action_items', 'notulensi_attachments', 'leaders')
        meetings_map = {m.pk: m for m in meetings}

    for a in page_obj:
        im = meetings_map.get(a.internal_meeting_id) if a.internal_meeting_id else None
        if im:
            a.internal_meeting = im

        att_list = []
        if im:
            if im.notulensi_file:
                att_list.append({'url': im.notulensi_file.url, 'name': 'Berkas Utama Notulensi'})
            for att in im.notulensi_attachments.all():
                att_list.append({'url': att.file.url, 'name': att.file_name or 'Dokumen / Foto Notulensi'})
        elif a.completed_file:
            att_list.append({'url': a.completed_file.url, 'name': 'Berkas Dokumen Hasil Agenda'})

        action_plan_list = []
        if im:
            for item in im.action_items.all():
                action_plan_list.append({
                    'id': item.pk,
                    'title': item.title,
                    'pic_id': item.pic_id or '',
                    'due_date': item.due_date.strftime('%Y-%m-%d') if item.due_date else ''
                })

        p_ids = []
        if im:
            p_ids = [p.pk for p in im.participants.all()]
        else:
            p_ids = [emp.pk for emp in a.assigned_employees.all()]

        notulensi_data = {
            'id': a.pk,
            'meeting_id': im.pk if im else (a.internal_meeting_id or ''),
            'meeting_number': im.meeting_number if (im and im.meeting_number) else a.title,
            'title': a.title,
            'scheduled_at': a.scheduled_at.strftime('%d %b %Y, %H:%M WIB') if a.scheduled_at else '-',
            'leader_names': im.leader_names_display if im else 'Pimpinan BAZNAS',
            'leader_id': (im.leader_id if (im and im.leader_id) else '') or '',
            'notulis_id': im.notulis_id if im else '',
            'notulis_name': im.notulis_name_display if im else '-',
            'meeting_type': im.meeting_type if im else 'khusus',
            'guest_names': (im.guest_names if im else '') or '',
            'status': a.status,
            'summary': (im.notulensi_summary if (im and im.notulensi_summary) else a.completed_notes) or '',
            'decision': (im.notulensi_decision if im else '') or '',
            'action_items_text': (im.notulensi_action_items if im else '') or '',
            'has_notulensi': bool((im and im.is_notulensi_completed) or a.completed_notes or a.completed_file),
            'participants': p_ids,
            'action_plan_items': action_plan_list,
            'attachments': att_list,
            'print_url': f'/rapat-internal/{im.pk}/print/' if im else '#'
        }
        a.notulensi_json = json.dumps(notulensi_data)

    from services.integrations.google_calendar_service import GoogleCalendarService
    google_cal_url = GoogleCalendarService.get_google_calendar_direct_url(request)

    agenda_dates_map = {}
    for ag in Agenda.objects.all().select_related('archive'):
        if ag.scheduled_at:
            local_dt = timezone.localtime(ag.scheduled_at)
            d_str = local_dt.strftime('%Y-%m-%d')
            t_str = local_dt.strftime('%H:%M WIB')
            icon = '👥' if ag.internal_meeting_id else '📅'
            b_text = f"{icon} {ag.title} ({t_str})"
            if d_str not in agenda_dates_map:
                agenda_dates_map[d_str] = []
            agenda_dates_map[d_str].append({
                'title': b_text,
                'type': 'meeting' if ag.internal_meeting_id else 'agenda',
                'url': f"/rapat-internal/{ag.internal_meeting_id}/" if ag.internal_meeting_id else f"/agenda/"
            })

    agenda_dates_json = json.dumps(agenda_dates_map)

    return render(request, 'agendas/list.html', {
        'page_obj': page_obj,
        'agendas': page_obj,
        'filters': {
            'status': status or '',
            'disposer': disposer or '',
            'terkait_user': terkait_user or '',
            'q': query or '',
        },
        'disposer_users': disposer_users,
        'terkait_users': terkait_users,
        'archive_types': Archive.TYPE_CHOICES if hasattr(Archive, 'TYPE_CHOICES') else [],
        'current_type': archive_type or '',
        'employees': Employee.objects.filter(is_active=True).order_by('full_name'),
        'google_calendar_direct_url': google_cal_url,
        'agenda_dates_json': agenda_dates_json,
        'can_manage_agenda': is_superadmin_or_kabid_4(request.user),
    })


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_complete(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    if request.method == 'POST':
        notes = request.POST.get('completed_notes', '')
        uploaded_file = request.FILES.get('completed_file')

        # 1. Update status Agenda
        agenda.is_completed = True
        agenda.status = 'selesai'
        agenda.completed_notes = notes
        if uploaded_file:
            agenda.completed_file = uploaded_file
        agenda.save()

        # 2. SINKRONISASI OTOMATIS KE SPPD TERKAIT -> SELESAI
        from sppd_service.models import SPPD
        if hasattr(agenda, 'sppd_ref') and agenda.sppd_ref:
            agenda.sppd_ref.status = 'selesai'
            if notes and not agenda.sppd_ref.report_notes:
                agenda.sppd_ref.report_notes = notes
            if uploaded_file and not agenda.sppd_ref.report_file:
                agenda.sppd_ref.report_file = uploaded_file
            agenda.sppd_ref.save()

        # 3. KELOLA STATUS ARSIP (Tetap 'proses' jika masih dalam alur SPPD)
        if agenda.archive:
            SPPD.objects.filter(
                disposition__archive=agenda.archive, 
                sppd_ref=agenda.sppd_ref if hasattr(agenda, 'sppd_ref') else None,
                status__in=['draft', 'disetujui', 'berlangsung']
            ).update(status='selesai')

            if uploaded_file and hasattr(agenda.archive, 'file_path') and not agenda.archive.file_path:
                agenda.archive.file_path = uploaded_file
                agenda.archive.save(update_fields=['file_path'])
        else:
            if uploaded_file or notes:
                new_archive = Archive.objects.create(
                    title=f"Notulensi: {agenda.title}",
                    description=notes or f"Laporan hasil pelaksanaan agenda {agenda.title}",
                    file_path=uploaded_file if hasattr(Archive, 'file_path') else None,
                    status='selesai',
                    created_by=request.user
                )
                agenda.archive = new_archive
                agenda.save(update_fields=['archive'])

        AuditService.log_action(request.user, f"Selesaikan Agenda: {agenda.title}", request)
        messages.success(request, f"Agenda '{agenda.title}' dan SPPD terkait berhasil diselesaikan!")
    return redirect('agendas:list')


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_cancel(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    if request.method == 'POST':
        agenda.status = 'dibatalkan'
        agenda.save()
        
        # Batalkan SPPD terkait jika agenda dibatalkan
        if hasattr(agenda, 'sppd_ref') and agenda.sppd_ref:
            agenda.sppd_ref.is_cancelled = True
            agenda.sppd_ref.status = 'dibatalkan'
            agenda.sppd_ref.save()

        AuditService.log_action(request.user, f"Batalkan Agenda: {agenda.title}", request)
        messages.success(request, f"Agenda '{agenda.title}' berhasil dibatalkan.")
    return redirect('agendas:list')


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_edit(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    if agenda.is_sppd_generated or agenda.sppd_ref:
        messages.warning(request, "Agenda yang terbit dari SPPD tidak dapat diedit secara manual di modul Agenda. Silakan lakukan pengubahan jadwal melalui modul SPPD.")
        return redirect('agendas:list')

    if request.method == 'POST':
        form = AgendaForm(request.POST, request.FILES, instance=agenda)
        if form.is_valid():
            title = request.POST.get('title')
            location = request.POST.get('location')
            description = request.POST.get('description')
            scheduled_at = request.POST.get('scheduled_at')
            archive_id = request.POST.get('archive')
            assigned_emp_ids = request.POST.getlist('assigned_to')
            wa_emp_ids = request.POST.getlist('wa_recipients')
            send_wa = request.POST.get('send_wa') == 'on'
            attachment = request.FILES.get('attachment')

            is_recurring = request.POST.get('is_recurring') == 'on'
            recurrence_type = request.POST.get('recurrence_type', 'none')
            recurrence_day = request.POST.get('recurrence_day') or None
            recurrence_end_date = request.POST.get('recurrence_end_date') or None
            wa_notification_timing = request.POST.get('wa_notification_timing', 'instant')

            archive = Archive.objects.filter(id=archive_id).first() if archive_id else agenda.archive

            old_scheduled_at = agenda.scheduled_at
            agenda = form.save(commit=False)
            agenda.title = title
            agenda.location = location
            agenda.description = description
            if scheduled_at:
                agenda.scheduled_at = scheduled_at
            agenda.archive = archive
            agenda.is_recurring = is_recurring
            agenda.recurrence_type = recurrence_type if is_recurring else 'none'
            agenda.recurrence_day = int(recurrence_day) if (is_recurring and recurrence_day is not None and str(recurrence_day).isdigit()) else None
            agenda.recurrence_end_date = recurrence_end_date if is_recurring else None
            agenda.wa_notification_timing = wa_notification_timing
            
            if old_scheduled_at and str(old_scheduled_at) != str(scheduled_at):
                agenda.status = 'diundur'
            
            if attachment:
                agenda.attachment = attachment
            agenda.save()

            assigned_emps = Employee.objects.filter(id__in=assigned_emp_ids)
            assigned_users = User.objects.filter(employee__in=assigned_emps)
            agenda.assigned_employees.set(assigned_emps)
            agenda.assigned_to.set(assigned_users)

            if send_wa:
                wa_emps = Employee.objects.filter(id__in=wa_emp_ids) if wa_emp_ids else assigned_emps
                
                file_link = ""
                if agenda.attachment:
                    file_link = f"\n📎 *Lampiran:* {request.build_absolute_uri(agenda.attachment.url)}"
                elif archive and hasattr(archive, 'file_path') and archive.file_path:
                    file_link = f"\n🔗 *Berkas:* {request.build_absolute_uri(archive.file_path.url)}"
                
                msg = (
                    f"Assalamu'alaikum Warahmatullahi Wabarakatuh,\n\n"
                    f"Yth. Bapak/Ibu Amil BAZNAS Kabupaten Tangerang,\n\n"
                    f"Pemberitahuan penyesuaian/perubahan jadwal agenda kegiatan dinas:\n\n"
                    f"• *Kegiatan:* {title}\n"
                    f"• *Jadwal Lama:* {old_scheduled_at.strftime('%d/%m/%Y %H:%M') if old_scheduled_at else '-'} WIB\n"
                    f"• *Jadwal Baru:* {scheduled_at} WIB\n"
                )
                if description:
                    msg += f"• *Keterangan:* {description}\n"
                msg += f"{file_link}\n\n" if file_link else "\n"
                msg += (
                    f"Mohon untuk dapat menyesuaikan kehadiran Anda sesuai dengan jadwal baru yang ditetapkan.\n\n"
                    f"Terima kasih atas perhatian dan pengertiannya.\n"
                    f"Wassalamu'alaikum Warahmatullahi Wabarakatuh."
                )
                
                user_map = {u.employee_id: u for u in User.objects.filter(employee__in=wa_emps)}
                for emp in wa_emps:
                    user = user_map.get(emp.pk)
                    WhatsAppService.send_notification(
                        user=user, 
                        message=msg, 
                        employee=emp, 
                        category='agenda', 
                        title="Perubahan Jadwal Agenda"
                    )
                
                agenda.notification_sent_at = timezone.now()
                agenda.save(update_fields=['notification_sent_at'])

            AuditService.log_action(request.user, f"Perbarui Agenda: {title}", request)
            messages.success(request, f"Agenda '{title}' berhasil diperbarui.")
            return redirect('agendas:list')
        else:
            messages.error(request, 'Data agenda tidak valid.')

    archives = Archive.objects.exclude(status='baru')
    employees = Employee.objects.filter(is_active=True).order_by('full_name')
    assigned_emp_ids = set(
        agenda.assigned_employees.values_list('id', flat=True)
    ) | set(
        agenda.assigned_to.filter(employee__isnull=False).values_list('employee_id', flat=True)
    )
    return render(request, 'agendas/create.html', {
        'archives': archives,
        'employees': employees,
        'agenda': agenda,
        'reschedule': True,
        'assigned_emp_ids': assigned_emp_ids,
    })


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_delete(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    if request.method == 'POST':
        title = agenda.title
        try:
            from .models import AgendaAttachment
            AgendaAttachment.objects.filter(agenda=agenda).delete()
        except Exception:
            pass
        agenda.delete()
        AuditService.log_action(request.user, f"Hapus Agenda: {title}", request)
        messages.success(request, f"Agenda '{title}' berhasil dihapus.")
    return redirect('agendas:list')


@login_required
def agenda_create(request):
    if not is_superadmin_or_kabid_4(request.user):
        messages.error(request, "Akses Ditolak: Hanya Superadmin dan Kabid IV yang berhak menambah agenda kegiatan.")
        return redirect('agendas:list')
    if request.method == 'POST':
        title = request.POST.get('title')
        location = request.POST.get('location')
        description = request.POST.get('description')
        scheduled_at = request.POST.get('scheduled_at')
        archive_id = request.POST.get('archive')
        assigned_emp_ids = request.POST.getlist('assigned_to')
        wa_emp_ids = request.POST.getlist('wa_recipients')
        send_wa = request.POST.get('send_wa') == 'on'
        attachment = request.FILES.get('attachment')
        
        is_recurring = request.POST.get('is_recurring') == 'on'
        recurrence_type = request.POST.get('recurrence_type', 'none')
        recurrence_day = request.POST.get('recurrence_day') or None
        recurrence_end_date = request.POST.get('recurrence_end_date') or None
        wa_notification_timing = request.POST.get('wa_notification_timing', 'instant')

        archive = Archive.objects.filter(id=archive_id).first() if archive_id else None
        
        agenda = Agenda.objects.create(
            title=title,
            location=location,
            description=description,
            scheduled_at=scheduled_at,
            archive=archive,
            attachment=attachment,
            created_by=request.user,
            is_recurring=is_recurring,
            recurrence_type=recurrence_type if is_recurring else 'none',
            recurrence_day=int(recurrence_day) if (is_recurring and recurrence_day is not None and str(recurrence_day).isdigit()) else None,
            recurrence_end_date=recurrence_end_date if is_recurring else None,
            wa_notification_timing=wa_notification_timing
        )
        
        assigned_emps = Employee.objects.filter(id__in=assigned_emp_ids)
        assigned_users = User.objects.filter(employee__in=assigned_emps)
        agenda.assigned_employees.set(assigned_emps)
        agenda.assigned_to.set(assigned_users)
        
        if send_wa:
            wa_emps = Employee.objects.filter(id__in=wa_emp_ids) if wa_emp_ids else assigned_emps
            
            file_link = ""
            if agenda.attachment:
                file_link = f"\n📎 *Lampiran:* {request.build_absolute_uri(agenda.attachment.url)}"
            elif archive and hasattr(archive, 'file_path') and archive.file_path:
                file_link = f"\n🔗 *Berkas:* {request.build_absolute_uri(archive.file_path.url)}"
            
            msg = (
                f"Assalamu'alaikum Warahmatullahi Wabarakatuh,\n\n"
                f"Yth. Bapak/Ibu Amil BAZNAS Kabupaten Tangerang,\n\n"
                f"Pemberitahuan agenda kegiatan dinas baru yang dijadwalkan:\n\n"
                f"• *Kegiatan:* {title}\n"
                f"• *Waktu:* {scheduled_at} WIB\n"
            )
            if description:
                msg += f"• *Keterangan:* {description}\n"
            msg += f"{file_link}\n\n" if file_link else "\n"
            msg += (
                f"Mohon untuk dapat bersiap dan menghadiri kegiatan sesuai waktu yang ditetapkan.\n\n"
                f"Terima kasih.\n"
                f"Wassalamu'alaikum Warahmatullahi Wabarakatuh."
            )
            
            user_map = {u.employee_id: u for u in User.objects.filter(employee__in=wa_emps)}
            for emp in wa_emps:
                user = user_map.get(emp.pk)
                WhatsAppService.send_notification(
                    user=user, 
                    message=msg, 
                    employee=emp, 
                    category='agenda', 
                    title="Agenda Baru"
                )
            
            agenda.notification_sent_at = timezone.now()
            agenda.save(update_fields=['notification_sent_at'])
            msg_notif = f"dan notifikasi WA terkirim ke {wa_emps.count()} pegawai"
        else:
            msg_notif = "tanpa notifikasi WA"

        AuditService.log_action(request.user, f"Buat Agenda Manual: {title}", request)
        messages.success(request, f"Agenda berhasil ditambahkan {msg_notif}.")
        return redirect('agendas:list')

        
    archive_param = request.GET.get('archive_id') or request.GET.get('archive')
    dispo_param = request.GET.get('disposition_id') or request.GET.get('disposition')
    
    selected_archive = None
    if archive_param and archive_param.isdigit():
        selected_archive = Archive.objects.filter(id=int(archive_param)).first()
    elif dispo_param and dispo_param.isdigit():
        from dispositions.models import Disposition
        dispo = Disposition.objects.filter(id=int(dispo_param)).first()
        if dispo and dispo.archive:
            selected_archive = dispo.archive

    archives = Archive.objects.exclude(status='baru')
    employees = Employee.objects.filter(is_active=True).order_by('full_name')
    return render(request, 'agendas/create.html', {
        'archives': archives,
        'employees': employees,
        'selected_archive': selected_archive,
    })


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_notify(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    if request.method == 'POST':
        archive = agenda.archive
        tgl_pelaksanaan = '—'
        if archive:
            dispo = archive.dispositions.first()
            if dispo and dispo.implementation_date:
                tgl_pelaksanaan = dispo.implementation_date.strftime('%d/%m/%Y')

        msg = (
            f"Assalamu'alaikum Warahmatullahi Wabarakatuh,\n\n"
            f"Yth. Bapak/Ibu Amil BAZNAS Kabupaten Tangerang,\n\n"
            f"Pengingat agenda kegiatan dinas BAZNAS Kabupaten Tangerang:\n\n"
            f"• *Kegiatan:* {agenda.title}\n"
            f"• *Waktu Agenda:* {agenda.scheduled_at.strftime('%d/%m/%Y %H:%M') if agenda.scheduled_at else '-'} WIB\n"
        )
        if tgl_pelaksanaan != '—':
            msg += f"• *Tanggal Pelaksanaan:* {tgl_pelaksanaan}\n"
        if agenda.description:
            msg += f"• *Keterangan:* {agenda.description}\n"
        if archive:
            msg += f"• *No. Arsip:* {archive.archive_number or '—'}\n"
            msg += f"• *Perihal Arsip:* {archive.title}\n"
            if hasattr(archive, 'file_path') and archive.file_path:
                msg += f"• *Link Berkas:* {request.build_absolute_uri(archive.file_path.url)}\n"
        if agenda.attachment:
            msg += f"• *Lampiran:* {request.build_absolute_uri(agenda.attachment.url)}\n"
        msg += (
            f"\nMohon untuk dapat bersiap dan menghadiri kegiatan tepat waktu.\n\n"
            f"Terima kasih.\n"
            f"Wassalamu'alaikum Warahmatullahi Wabarakatuh."
        )

        # Target penerima notifikasi WA: seluruh Pegawai Ditugaskan & User terkait
        target_employees = set(agenda.assigned_employees.all())
        for u in agenda.assigned_to.all():
            if hasattr(u, 'employee') and u.employee:
                target_employees.add(u.employee)

        if archive:
            dispo_emps = Employee.objects.filter(received_dispositions__archive=archive)
            for de in dispo_emps:
                target_employees.add(de)

        if not target_employees:
            messages.warning(request, f"Tidak ada pegawai yang ditugaskan untuk agenda '{agenda.title}'. Tambahkan pegawai melalui edit agenda.")
            return redirect('agendas:list')

        sent = 0
        user_map = {u.employee_id: u for u in User.objects.filter(employee__in=target_employees)}
        for emp in target_employees:
            user = user_map.get(emp.pk)
            if WhatsAppService.send_notification(user=user, message=msg, employee=emp, category='agenda', title="Pengingat Agenda"):
                sent += 1

        agenda.notification_sent_at = timezone.now()
        agenda.save(update_fields=['notification_sent_at'])

        AuditService.log_action(request.user, f"Kirim Notifikasi Pengingat Agenda: {agenda.title}", request)
        messages.success(request, f"Notifikasi WA dikirim ke {sent} dari {len(target_employees)} pegawai ditugaskan untuk agenda '{agenda.title}'.")
    return redirect('agendas:list')


@login_required
def agenda_events(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    disposer = request.GET.get('disposer')
    terkait_user = request.GET.get('terkait_user')
    
    events = Agenda.objects.select_related('archive', 'created_by').prefetch_related('assigned_to', 'assigned_employees').all()

    if start and end:
        try:
            from django.utils.dateparse import parse_datetime
            start_dt = parse_datetime(start)
            end_dt = parse_datetime(end)
            if start_dt and end_dt:
                events = events.filter(scheduled_at__gte=start_dt, scheduled_at__lte=end_dt)
        except Exception:
            pass

    if disposer:
        events = events.filter(archive__dispositions__sender_id=disposer)
    if terkait_user:
        events = events.filter(
            Q(archive__dispositions__forwarded_to__user_account__id=terkait_user) |
            Q(archive__dispositions__sppd_list__assigned_employees__user_account__id=terkait_user) |
            Q(assigned_to__id=terkait_user)
        )
        
    data = []
    for event in events:
        if event.status == 'selesai':
            bg_color = '#6B7280'
            text_color = '#FFFFFF'
        elif event.status == 'dibatalkan':
            bg_color = '#EF4444'
            text_color = '#FFFFFF'
        elif event.status == 'diundur':
            bg_color = '#F59E0B'
            text_color = '#1F2937'
        else:
            bg_color = '#0E9F6E'
            text_color = '#FFFFFF'

        if event.archive and event.status == 'terjadwal':
            if event.archive.status == 'baru':
                bg_color = '#F59E0B'
            elif event.archive.status == 'proses':
                bg_color = '#3B82F6'

        event_url = '#'
        if event.completed_file:
            event_url = event.completed_file.url
        elif event.archive:
            event_url = f'/archives/{event.archive.pk}/'

        src_type = 'sppd' if (event.is_sppd_generated or event.sppd_ref) else ('meeting' if event.internal_meeting_id else 'agenda')
        local_dt = timezone.localtime(event.scheduled_at) if event.scheduled_at else None
        date_str = local_dt.strftime('%Y-%m-%d') if local_dt else None
        time_str = local_dt.strftime('%H:%M WIB') if local_dt else ''
        display_title = f"{event.title} ({time_str})" if time_str else event.title

        data.append({
            'id': event.id,
            'title': display_title,
            'start': date_str,
            'allDay': True,
            'display': 'block',
            'backgroundColor': bg_color,
            'borderColor': bg_color,
            'textColor': text_color,
            'url': event_url,
            'extendedProps': {
                'location': event.location or 'Kantor BAZNAS',
                'assigned_to': event.assigned_names_display or 'Internal',
                'source_type': src_type,
                'status': event.get_status_display()
            }
        })

    from services.integrations.google_calendar_service import GoogleCalendarService
    holiday_events = GoogleCalendarService.get_national_holidays_events(start_year=2025, end_year=2027)
    data.extend(holiday_events)

    return JsonResponse(data, safe=False)


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_generate_sppd(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    scheduled_date = agenda.scheduled_at.date() if agenda.scheduled_at else timezone.now().date()
    
    archive = agenda.archive
    if not archive:
        from archives.models import Category
        category = Category.objects.filter(name__icontains='surat').first() or Category.objects.first()
        
        from services.archives.numbering_service import NumberingService
        arc_num = NumberingService.generate_number('surat_keluar')
        
        archive = Archive.objects.create(
            archive_number=arc_num,
            title=agenda.title,
            archive_type='surat_keluar',
            category=category,
            uploaded_by=request.user,
            sender='BAZNAS Kabupaten Tangerang',
            receiver=agenda.location or 'Lokasi Penugasan Agenda',
            letter_date=scheduled_date,
            received_date=scheduled_date,
            description=f"Dokumen Penugasan Internal dari Agenda: {agenda.title}",
            status='sudah_ditugaskan'
        )
        agenda.archive = archive
        agenda.save(update_fields=['archive', 'updated_at'])
    else:
        if not archive.letter_date:
            archive.letter_date = scheduled_date
            archive.received_date = scheduled_date
            archive.save(update_fields=['letter_date', 'received_date', 'updated_at'])

    dispo = archive.dispositions.first()
    if not dispo:
        dispo = Disposition.objects.create(
            archive=archive,
            sender=request.user,
            priority='segera',
            inst_laporkan=True,
            inst_koordinasikan=True,
            instructions='Laporkan Hasil Pelaksanaan, Koordinasikan / Tindak Lanjut',
            note=f"Disposisi Penugasan Agenda: {agenda.title}",
            status='terisi'
        )
        if agenda.assigned_employees.exists():
            dispo.forwarded_to.set(agenda.assigned_employees.all())
    else:
        dispo.priority = 'segera'
        dispo.inst_laporkan = True
        dispo.inst_koordinasikan = True
        if agenda.assigned_employees.exists() and not dispo.forwarded_to.exists():
            dispo.forwarded_to.set(agenda.assigned_employees.all())
        dispo.save()

    from surat_tugas.models import SuratTugas
    st = SuratTugas.objects.filter(disposition=dispo).first()
    if not st:
        st = SuratTugas.objects.create(
            disposition=dispo,
            tentang=agenda.title,
            lokasi_tujuan=agenda.location or 'Kabupaten Tangerang',
            tanggal_mulai=scheduled_date,
            created_by=request.user
        )
        if agenda.assigned_employees.exists():
            st.pegawai_ditugaskan.set(agenda.assigned_employees.all())
    else:
        if agenda.assigned_employees.exists() and not st.pegawai_ditugaskan.exists():
            st.pegawai_ditugaskan.set(agenda.assigned_employees.all())

    AuditService.log_action(request.user, f"Terbitkan SPPD dari Agenda: {agenda.title}", request)
    messages.success(request, f"Penugasan untuk Agenda '{agenda.title}' berhasil diselaraskan. Silakan lengkapi detail SPPD.")
    return redirect('sppd_service:create', dispo_pk=dispo.pk)


@login_required
def agenda_upload_notulensi(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    
    if request.method == 'POST':
        summary = request.POST.get('notulensi_summary') or request.POST.get('completed_notes', '').strip()
        decision = request.POST.get('notulensi_decision', '').strip()
        action_items = request.POST.get('notulensi_action_items', '').strip()
        notulis_id = request.POST.get('notulis')
        status_val = request.POST.get('status', 'selesai')
        participant_ids = request.POST.getlist('participants')
        
        files = request.FILES.getlist('notulensi_files') or request.FILES.getlist('completed_files') or request.FILES.getlist('completed_file') or request.FILES.getlist('notulensi_file')
        file_labels = request.POST.getlist('notulensi_file_labels[]')
        single_file = request.FILES.get('notulensi_file') or (files[0] if files else None)

        notes_combined = summary
        if decision:
            notes_combined += f"\n\nKeputusan:\n{decision}"
        if action_items:
            notes_combined += f"\n\nRencana Tindak Lanjut:\n{action_items}"

        agenda.completed_notes = notes_combined
            
        if single_file:
            agenda.completed_file = single_file
        if files:
            from .models import AgendaAttachment
            for idx, uploaded_f in enumerate(files):
                label = file_labels[idx] if idx < len(file_labels) and file_labels[idx].strip() else f"Dokumentasi #{idx+1} - {agenda.title}"
                AgendaAttachment.objects.create(
                    agenda=agenda,
                    file=uploaded_f,
                    description=label
                )

        agenda.is_completed = True
        agenda.status = status_val if status_val in ['terjadwal', 'berlangsung', 'selesai', 'dibatalkan'] else 'selesai'
        agenda.save()

        # Update or Create synced InternalMeeting & Action Items
        from internal_meetings.models import InternalMeeting, MeetingActionItem, MeetingNotulensiAttachment
        from users.models import Employee

        meeting = None
        if agenda.internal_meeting_id:
            meeting = InternalMeeting.objects.filter(pk=agenda.internal_meeting_id).first()

        notulis_emp = Employee.objects.filter(pk=int(notulis_id)).first() if (notulis_id and str(notulis_id).isdigit()) else None
        leader_id = request.POST.get('leader')
        leader_emp = Employee.objects.filter(pk=int(leader_id)).first() if (leader_id and str(leader_id).isdigit()) else None

        guest_names_val = request.POST.get('guest_names', '').strip()
        meeting_type_val = request.POST.get('meeting_type', '')

        if not meeting and (summary or decision or request.POST.getlist('action_title[]') or guest_names_val):
            try:
                meeting = InternalMeeting.objects.create(
                    title=agenda.title,
                    meeting_type=meeting_type_val if meeting_type_val else ('audiensi' if ('audiensi' in agenda.title.lower() or 'tamu' in agenda.title.lower()) else 'khusus'),
                    scheduled_at=agenda.scheduled_at,
                    location=agenda.location or 'Kantor BAZNAS',
                    agenda_topics=agenda.description or agenda.title,
                    status=agenda.status,
                    created_by=request.user,
                    notulensi_summary=summary,
                    notulensi_decision=decision,
                    notulensi_action_items=action_items,
                    guest_names=guest_names_val,
                    leader=leader_emp,
                    notulis=notulis_emp,
                    notulensi_created_at=timezone.now()
                )
                if leader_emp:
                    meeting.leaders.set([leader_emp])
                elif agenda.assigned_employees.exists():
                    meeting.leaders.set(agenda.assigned_employees.all())

                agenda.description = f"InternalMeetingID:{meeting.pk}\n\n{agenda.description or ''}"
                agenda.save(update_fields=['description'])
            except Exception as e_mtg:
                print("Error creating InternalMeeting for agenda:", e_mtg)

        if meeting:
            try:
                meeting.notulensi_summary = summary
                meeting.notulensi_decision = decision
                meeting.notulensi_action_items = action_items
                meeting.guest_names = guest_names_val
                if meeting_type_val:
                    meeting.meeting_type = meeting_type_val
                if leader_emp:
                    meeting.leader = leader_emp
                    meeting.leaders.set([leader_emp])
                elif not meeting.leaders.exists() and agenda.assigned_employees.exists():
                    meeting.leaders.set(agenda.assigned_employees.all())

                if notulis_emp:
                    meeting.notulis = notulis_emp
                if participant_ids:
                    p_emps = Employee.objects.filter(pk__in=participant_ids)
                    meeting.participants.set(p_emps)
                if single_file:
                    meeting.notulensi_file = single_file
                meeting.status = agenda.status
                meeting.notulensi_created_at = timezone.now()
                meeting.save()

                if files:
                    for idx, uploaded_f in enumerate(files):
                        label = file_labels[idx] if idx < len(file_labels) and file_labels[idx].strip() else f"Lampiran #{idx+1}"
                        MeetingNotulensiAttachment.objects.create(
                            meeting=meeting,
                            file=uploaded_f,
                            file_name=label
                        )

                # Process Dynamic Action Plan Items
                action_titles = request.POST.getlist('action_title[]')
                action_pics = request.POST.getlist('action_pic[]')
                action_due_dates = request.POST.getlist('action_due_date[]')
                action_ids = request.POST.getlist('action_id[]')

                processed_ids = []
                for i, title in enumerate(action_titles):
                    title_str = title.strip()
                    if not title_str:
                        continue

                    item_id = action_ids[i] if i < len(action_ids) else None
                    pic_id = action_pics[i] if i < len(action_pics) and action_pics[i] else None
                    due_date_val = action_due_dates[i] if i < len(action_due_dates) and action_due_dates[i] else None
                    pic_obj = Employee.objects.filter(pk=pic_id).first() if pic_id else None

                    if item_id and item_id.isdigit():
                        item = MeetingActionItem.objects.filter(pk=int(item_id), meeting=meeting).first()
                        if item:
                            item.title = title_str
                            item.pic = pic_obj
                            if due_date_val:
                                item.due_date = due_date_val
                            item.save()
                            processed_ids.append(item.pk)
                            continue

                    new_item = MeetingActionItem.objects.create(
                        meeting=meeting,
                        title=title_str,
                        pic=pic_obj,
                        due_date=due_date_val if due_date_val else None,
                        is_tracked=True
                    )
                    processed_ids.append(new_item.pk)
            except Exception as err:
                print("Error updating InternalMeeting from Agenda notulensi:", err)

        agenda.status = 'selesai'
        agenda.save()

        # -------------------------------------------------------------
        # OTOMATIS SINKRON SIKLUS SPPD TAHAP 1 -> SELESAI
        # -------------------------------------------------------------
        from sppd_service.models import SPPD
        if hasattr(agenda, 'sppd_ref') and agenda.sppd_ref:
            agenda.sppd_ref.status = 'selesai'
            if notes_combined and not agenda.sppd_ref.report_notes:
                agenda.sppd_ref.report_notes = notes_combined
            if files and not agenda.sppd_ref.report_file:
                agenda.sppd_ref.report_file = files[0]
            agenda.sppd_ref.save()

        try:
            archive = agenda.archive
            dispo = archive.dispositions.first() if archive else None
            
            if archive:
                SPPD.objects.filter(
                    disposition__archive=archive, 
                    status__in=['draft', 'disetujui', 'berlangsung']
                ).update(status='selesai')

            if dispo:
                from reports.models import Report, ReportAttachment
                report, created = Report.objects.get_or_create(
                    disposition=dispo,
                    defaults={
                        'title': f"Laporan Hasil: {agenda.title}",
                        'content': notes_combined or f"Kegiatan '{agenda.title}' telah selesai dilaksanakan.",
                        'created_by': request.user
                    }
                )
                if not created:
                    if notes_combined:
                        report.content = notes_combined
                    report.save()

                if files:
                    report.file = files[0]
                    report.save(update_fields=['file'])
                    for idx, uploaded_f in enumerate(files):
                        ReportAttachment.objects.create(
                            report=report,
                            file=uploaded_f,
                            description=f"Dokumentasi #{idx+1} Agenda: {agenda.title}"
                        )
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Failed to auto sync Agenda Notulensi to Report: %s", e)

        AuditService.log_action(request.user, f"Upload Notulensi Agenda: {agenda.title}", request)
        messages.success(request, f"Notulensi & Hasil Agenda '{agenda.title}' berhasil disimpan.")
        return redirect('agendas:list')

    return redirect('agendas:list')


@login_required
@staff_or_kabid_or_pimpinan_required
def agenda_complete(request, pk):
    agenda = get_object_or_404(Agenda, pk=pk)
    
    # Check if this is an SPPD agenda or has an archive / disposition
    archive = agenda.archive
    if not archive and hasattr(agenda, 'sppd_ref') and agenda.sppd_ref:
        archive = agenda.sppd_ref.disposition.archive if (agenda.sppd_ref and agenda.sppd_ref.disposition) else None

    if archive:
        dispo = archive.dispositions.first()
        if dispo:
            # Direct user to Report Input page for complete integration
            from reports.models import Report
            existing_report = Report.objects.filter(disposition=dispo).first()
            if existing_report:
                return redirect('reports:detail', pk=existing_report.pk)
            return redirect('reports:create', dispo_pk=dispo.pk)

    # Standard completion if no disposition attached
    agenda.is_completed = True
    agenda.status = 'selesai'
    if request.method == 'POST':
        notes = request.POST.get('completed_notes', '').strip()
        if notes:
            agenda.completed_notes = notes
    agenda.save()

    AuditService.log_action(request.user, f"Selesaikan Agenda: {agenda.title}", request)
    messages.success(request, f"Agenda '{agenda.title}' berhasil diselesaikan.")
    return redirect('agendas:list')


@login_required
def agenda_ical_feed(request):
    """
    Ekspor Feed iCal (.ics) agar user dapat berlangganan kalender SIMAP BAZNAS 
    secara otomatis di Google Calendar, Apple Calendar, atau Outlook.
    """
    agendas = Agenda.objects.filter(status__in=['terjadwal', 'selesai', 'diundur']).select_related('archive')
    
    ics_lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//BAZNAS Kabupaten Tangerang//SIMAP Agenda//ID",
        "X-WR-CALNAME:Agenda SIMAP BAZNAS",
        "X-WR-TIMEZONE:Asia/Jakarta",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH"
    ]
    
    for a in agendas:
        if not a.scheduled_at:
            continue
        dt_start = a.scheduled_at.strftime('%Y%m%dT%H%M%SZ')
        dt_stamp = a.created_at.strftime('%Y%m%dT%H%M%SZ') if a.created_at else dt_start
        summary = (a.title or '').replace('\n', ' ').replace(',', '\\,')
        location = (a.location or 'Kantor BAZNAS').replace('\n', ' ').replace(',', '\\,')
        description = (a.description or a.title or '').replace('\n', ' ').replace(',', '\\,')
        
        ics_lines.extend([
            "BEGIN:VEVENT",
            f"UID:simap-agenda-{a.id}@baznaskabtangerang.id",
            f"DTSTAMP:{dt_stamp}",
            f"DTSTART:{dt_start}",
            f"SUMMARY:{summary}",
            f"LOCATION:{location}",
            f"DESCRIPTION:{description}",
            "STATUS:CONFIRMED",
            "END:VEVENT"
        ])
        
    ics_lines.append("END:VCALENDAR")
    ics_content = "\r\n".join(ics_lines)
    
    response = HttpResponse(ics_content, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="agenda_simap_baznas.ics"'
    return response