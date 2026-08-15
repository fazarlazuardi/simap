import os
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.conf import settings

# Modul OpenPyXL untuk Format Excel Standar Lembaga
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage

from archives.models import Archive
from dispositions.models import Disposition
from users.models import User, SystemSetting
from .models import Report, ReportAttachment, MonthlyBackup
from services.archives.numbering_service import NumberingService
from services.audit_logs.audit_service import AuditService
from services.integrations.gateway_service import GoogleIntegrationService
from users.decorators import sdm_required


def _get_logo_path():
    """Helper untuk mencari keberadaan file Logo BAZNAS di berbagai direktori static."""
    possible_names = [
        'logo-baznas.png', 'logo_baznas.png', 'logo.png',
        'logo-baznas.jpg', 'logo_baznas.jpg', 'logo.jpg',
        'logo-baznas.jpeg', 'logo_baznas.jpeg', 'logo.jpeg',
        'logo-baznas-kab-tangerang.png', 'logo_baznas_kab_tangerang.png'
    ]
    
    search_dirs = [
        os.path.join(settings.BASE_DIR, 'static', 'img'),
        os.path.join(settings.BASE_DIR, 'static', 'images'),
        os.path.join(settings.BASE_DIR, 'static'),
        os.path.join(settings.BASE_DIR, 'apps', 'static', 'img'),
        os.path.join(settings.BASE_DIR, 'apps', 'static', 'images'),
        os.path.join(settings.BASE_DIR, 'apps', 'static'),
    ]
    
    if hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT:
        search_dirs.append(settings.STATIC_ROOT)
        search_dirs.append(os.path.join(settings.STATIC_ROOT, 'img'))
        search_dirs.append(os.path.join(settings.STATIC_ROOT, 'images'))
        
    if hasattr(settings, 'STATICFILES_DIRS') and settings.STATICFILES_DIRS:
        for d in settings.STATICFILES_DIRS:
            search_dirs.append(str(d))
            search_dirs.append(os.path.join(str(d), 'img'))
            search_dirs.append(os.path.join(str(d), 'images'))

    for d in search_dirs:
        if os.path.exists(d):
            for name in possible_names:
                full_p = os.path.join(d, name)
                if os.path.exists(full_p):
                    return full_p
    return None


@login_required
def report_index(request):
    """
    Traceability Table for documents dengan Ekspor Excel Standar Resmi BAZNAS.
    """
    query = request.GET.get('q')
    pj = request.GET.get('pj')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    archive_type = request.GET.get('type')

    archives = Archive.objects.select_related('category')\
        .prefetch_related(
            Prefetch(
                'dispositions',
                queryset=Disposition.objects.select_related('sender').prefetch_related('forwarded_to', 'waka_forwarded_to', 'report')
            ),
            'agendas'
        ).all().order_by('-updated_at')

    # Filter khusus Waka II, Kabid II & Staf Bidang II / POV Waka II
    active_pov = request.session.get('active_pov')
    is_waka_or_kabid_2 = active_pov in ['waka_2', 'kabid_2'] or getattr(request.user, 'is_waka_2', False) or getattr(request.user, 'is_kabid_2', False)
    if not is_waka_or_kabid_2:
        emp = getattr(request.user, 'employee', None)
        if emp and emp.dept_relation and ('pendistribusian' in emp.dept_relation.name.lower() or 'bidang ii' in emp.dept_relation.name.lower() or 'bidang 2' in emp.dept_relation.name.lower()):
            is_waka_or_kabid_2 = True

    if is_waka_or_kabid_2:
        from services.workflows.workflow_engine import WorkflowEngine
        archives = archives.filter(
            Q(verified_by_kabid=True) | ~Q(status='baru')
        )
        bantuan_ids = [arc.id for arc in archives if WorkflowEngine.is_bantuan(arc)]
        archives = archives.filter(id__in=bantuan_ids)

    if query:
        archives = archives.filter(
            Q(archive_number__icontains=query) |
            Q(title__icontains=query)
        )
    if pj:
        archives = archives.filter(
            Q(dispositions__forwarded_to__user_account__id=pj) |
            Q(dispositions__waka_forwarded_to__user_account__id=pj)
        ).distinct()

    if status:
        archives = archives.filter(status=status)
    if date_from:
        archives = archives.filter(updated_at__date__gte=date_from)
    if date_to:
        archives = archives.filter(updated_at__date__lte=date_to)
    if archive_type:
        archives = archives.filter(archive_type=archive_type)

    export = request.GET.get('export')
    if export:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trace Laporan Arsip"
        
        # Tampilkan Garis Tabel Bawaan
        ws.views.sheetView[0].showGridLines = True

        # Ketinggian Baris Area Kop Surat
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 10

        # --- 1. ATUR LEBAR KOLOM A KHUSUS UNTUK LOGO ---
        ws.column_dimensions['A'].width = 16

        # Container Logo BAZNAS (A2:A4)
        ws.merge_cells('A2:A4')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        logo_path = _get_logo_path()
        if logo_path:
            try:
                img = OpenPyXLImage(logo_path)
                # Height 0.66" (~63px), Width 1.06" (~102px)
                img.height = 63
                img.width = 102
                ws.add_image(img, 'A2')
            except Exception:
                pass

        # --- 2. TEKS KOP SURAT (MERGE B2:E2, B3:E3, B4:E4 -> MENYESUAIKAN PANJANG TEKS) ---
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        ws.merge_cells('B4:E4')

        cell_title1 = ws['B2']
        cell_title1.value = "BADAN AMIL ZAKAT NASIONAL (BAZNAS)"
        cell_title1.font = Font(name='Calibri', size=14, bold=True, color='006633')
        cell_title1.alignment = Alignment(horizontal='left', vertical='center')

        cell_title2 = ws['B3']
        cell_title2.value = "KABUPATEN TANGERANG"
        cell_title2.font = Font(name='Calibri', size=12, bold=True, color='222222')
        cell_title2.alignment = Alignment(horizontal='left', vertical='center')

        cell_title3 = ws['B4']
        cell_title3.value = f"LAPORAN TRACEABILITY PENGELOLAAN ARSIP & DISPOSISI (Dicetak: {timezone.now().strftime('%d %B %Y, %H:%M')} WIB)"
        cell_title3.font = Font(name='Calibri', size=9.5, italic=True, color='555555')
        cell_title3.alignment = Alignment(horizontal='left', vertical='center')

        # --- 3. ALL BORDER UNTUK BINGKAI KOP SURAT (A2:E4) & GARIS GANDA BARIS 5 ---
        black_thin = Side(style='thin', color='000000')
        solid_border = Border(left=black_thin, right=black_thin, top=black_thin, bottom=black_thin)
        double_bottom_side = Side(style='double', color='000000')

        # All Border khusus area Kop Surat A2:E4
        for r in range(2, 5):
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = solid_border

        # Garis pembatas ganda di bawah area kop (A5:E5)
        for col_idx in range(1, 6):
            ws.cell(row=5, column=col_idx).border = Border(bottom=double_bottom_side)

        # --- 4. HEADER TABEL DATA ---
        headers = [
            'NO.', 'NO. DOKUMEN', 'PENGIRIM', 'NAMA / PERIHAL DOKUMEN', 
            'NO. DISPOSISI', 'PENANGGUNG JAWAB AKTIF', 'TGL AGENDA', 
            'NO. SPPD / ST', 'NO. LAPORAN', 'STATUS WORKFLOW'
        ]

        start_row = 6
        ws.row_dimensions[start_row].height = 28

        fill_header = PatternFill(start_color='006633', end_color='006633', fill_type='solid')
        font_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header_text)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_header
            cell.border = solid_border

        # --- 5. BARIS DATA TABEL (ALL BORDER) ---
        current_row = start_row + 1
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        font_data = Font(name='Calibri', size=9.5)

        data_rows_values = []

        for idx, arc in enumerate(archives, 1):
            dispo = arc.latest_dispo
            pj_list = arc.current_assignee_names
            
            tgl_val = arc.latest_agenda_date
            tgl_agenda = tgl_val.strftime('%d/%m/%Y') if tgl_val else '-'
            
            sppd_obj = arc.latest_sppd
            st_obj = arc.latest_st
            sppd_no = sppd_obj.sppd_number if sppd_obj else (st_obj.nomor_surat if st_obj else '-')
            
            report_obj = arc.latest_report
            report_no = report_obj.report_number if report_obj else '-'
            
            status_str = arc.activity_name if arc.activity_name else arc.get_status_display()

            row_values = [
                idx,
                arc.archive_number or '(DRAFT)',
                arc.sender or '-',
                arc.title or '-',
                dispo.disposition_number or (f'DISP-{dispo.id}' if dispo else '-'),
                pj_list or '-',
                tgl_agenda,
                sppd_no,
                report_no,
                status_str.upper() if status_str else '-'
            ]
            
            data_rows_values.append(row_values)

            ws.row_dimensions[current_row].height = 24
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = solid_border

                if col_idx in [1, 2, 5, 7, 8, 9, 10]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

            current_row += 1

        # --- 6. DYNAMIC AUTO-FIT LEBAR KOLOM ---
        min_widths = {
            1: 16,   # NO. & CONTAINER LOGO
            2: 24,   # NO. DOKUMEN
            3: 22,   # PENGIRIM
            4: 38,   # NAMA DOKUMEN
            5: 18,   # NO. DISPOSISI
            6: 28,   # PENANGGUNG JAWAB
            7: 15,   # TGL AGENDA
            8: 20,   # NO. SPPD / ST
            9: 18,   # NO. LAPORAN
            10: 34   # STATUS WORKFLOW
        }

        for col_idx, header_text in enumerate(headers, 1):
            max_len = len(str(header_text))
            for r_val in data_rows_values:
                val_str = str(r_val[col_idx - 1]) if r_val[col_idx - 1] is not None else ''
                if len(val_str) > max_len:
                    max_len = len(val_str)

            col_letter = get_column_letter(col_idx)
            calculated_width = max(max_len + 4, min_widths.get(col_idx, 12))
            ws.column_dimensions[col_letter].width = calculated_width

        AuditService.log_action(request.user, "Ekspor Trace Laporan ke Excel Standar BAZNAS", request)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Trace_Laporan_BAZNAS_KabTangerang_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    users = User.objects.filter(is_active_account=True).order_by('username')
    drive_backup_enabled = SystemSetting.get_value('DRIVE_BACKUP_ENABLED', 'on') == 'on'

    per_page = int(request.GET.get('per_page', 500))
    paginator = Paginator(archives, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    now_year = timezone.now().year
    backup_years = range(now_year - 5, now_year + 2)
    existing_backups = MonthlyBackup.objects.all()[:10]

    return render(request, 'reports/index.html', {
        'page_obj': page_obj,
        'archives': page_obj,
        'q': query,
        'pj': pj or '',
        'status_filter': status or '',
        'date_from': date_from or '',
        'date_to': date_to or '',
        'users': users,
        'status_choices': Archive.STATUS_CHOICES,
        'drive_backup_enabled': drive_backup_enabled,
        'archive_types': Archive.TYPE_CHOICES,
        'current_type': archive_type or '',
        'backup_years': backup_years,
        'now_year': now_year,
        'existing_backups': existing_backups,
    })

@login_required
@sdm_required
def report_create(request, dispo_pk):
    """
    Input final report for a disposition.
    """
    dispo = get_object_or_404(Disposition, pk=dispo_pk)
    if request.method == 'POST':
        report_number = request.POST.get('report_number')
        title = request.POST.get('title')
        content = request.POST.get('content')
        
        report = Report.objects.create(
            disposition=dispo,
            report_number=report_number,
            title=title,
            content=content,
            created_by=request.user
        )
        
        # Robust multi attachment retrieval (supports all form field naming conventions)
        files = (
            request.FILES.getlist('attachment_file') or 
            request.FILES.getlist('attachment_file[]') or 
            request.FILES.getlist('file') or 
            request.FILES.getlist('file[]') or 
            request.FILES.getlist('attachment') or 
            request.FILES.getlist('attachments')
        )
        if not files:
            single_f = request.FILES.get('attachment_file') or request.FILES.get('file') or request.FILES.get('attachment')
            if single_f:
                files = [single_f]

        descriptions = (
            request.POST.getlist('attachment_description') or 
            request.POST.getlist('attachment_description[]') or 
            request.POST.getlist('description')
        )

        for i, f in enumerate(files):
            desc = descriptions[i] if i < len(descriptions) else ''
            att = ReportAttachment.objects.create(
                report=report,
                file=f,
                description=desc or f"Lampiran Berkas #{i+1}"
            )
            if i == 0:
                report.file = f
                report.save(update_fields=['file'])
        
        # Finalize the archive status
        archive = dispo.archive
        if archive:
            archive.status = 'selesai'
            archive.save(update_fields=['status', 'updated_at'])

            # Otomatis selesaikan SPPD & Agenda terkait
            from sppd_service.models import SPPD
            from agendas.models import Agenda
            sppds = SPPD.objects.filter(Q(disposition=dispo) | Q(disposition__archive=archive))
            sppds.update(status='selesai')

            agendas = Agenda.objects.filter(archive=archive)
            agendas.update(is_completed=True, status='selesai')
            if files:
                sppds.filter(report_file='').update(report_file=files[0])
            for s in sppds:
                if content and not s.report_notes:
                    s.report_notes = content
                    s.save(update_fields=['report_notes'])

            # Otomatis selesaikan Agenda Kerja terkait
            from agendas.models import Agenda
            agendas = Agenda.objects.filter(archive=archive)
            for ag in agendas:
                ag.is_completed = True
                ag.status = 'selesai'
                if not ag.completed_notes:
                    ag.completed_notes = f"Laporan Hasil Penanganan ({report_number}): {title}\n{content}"
                if files and not ag.completed_file:
                    ag.completed_file = files[0]
                ag.save()
        
        AuditService.log_action(request.user, f"Input Laporan Hasil Selesai: {report_number}", request)
        messages.success(request, f"Laporan Hasil '{report_number}' dengan {len(files)} berkas lampiran berhasil disimpan.")
        return redirect('reports:detail', pk=report.pk)
        
    report_type = request.GET.get('type') or request.GET.get('st_type')
    default_title = ""
    default_content = ""
    if report_type == 'penyaluran' and dispo and dispo.archive:
        default_title = f"Laporan Hasil Penyaluran Bantuan Mustahik: {dispo.archive.title}"
        default_content = f"Telah dilaksanakan penyaluran / pentasyarufan bantuan BAZNAS secara langsung (pentasyarufan direct) untuk permohonan '{dispo.archive.title}'."

    default_report_number = NumberingService.get_default_number('report')
    return render(request, 'reports/create.html', {
        'dispo': dispo,
        'default_report_number': default_report_number,
        'default_title': default_title,
        'default_content': default_content,
        'is_penyaluran_direct': (report_type == 'penyaluran'),
    })

@login_required
@sdm_required
def report_edit(request, pk):
    """
    Edit an existing report.
    """
    report = get_object_or_404(Report, pk=pk)
    
    if request.method == 'POST':
        report.report_number = request.POST.get('report_number')
        report.title = request.POST.get('title')
        report.content = request.POST.get('content')
        report.save()
        
        # Handle deletion of old attachments
        delete_ids = request.POST.getlist('delete_attachments')
        if delete_ids:
            ReportAttachment.objects.filter(report=report, id__in=delete_ids).delete()

        # Robust multi attachment retrieval
        files = (
            request.FILES.getlist('attachment_file') or 
            request.FILES.getlist('attachment_file[]') or 
            request.FILES.getlist('file') or 
            request.FILES.getlist('file[]') or 
            request.FILES.getlist('attachment') or 
            request.FILES.getlist('attachments')
        )
        if not files:
            single_f = request.FILES.get('attachment_file') or request.FILES.get('file') or request.FILES.get('attachment')
            if single_f:
                files = [single_f]

        descriptions = (
            request.POST.getlist('attachment_description') or 
            request.POST.getlist('attachment_description[]') or 
            request.POST.getlist('description')
        )

        for i, f in enumerate(files):
            desc = descriptions[i] if i < len(descriptions) else ''
            att = ReportAttachment.objects.create(
                report=report,
                file=f,
                description=desc or f"Lampiran Berkas #{i+1}"
            )
            if i == 0 and not report.file:
                report.file = f
                report.save(update_fields=['file'])
        
        # Sync changes to linked Archive, SPPD, and Agenda
        if report.disposition and report.disposition.archive:
            archive = report.disposition.archive
            archive.status = 'selesai'
            archive.save(update_fields=['status', 'updated_at'])

            from sppd_service.models import SPPD
            from agendas.models import Agenda
            SPPD.objects.filter(Q(disposition=report.disposition) | Q(disposition__archive=archive)).update(status='selesai')
            agendas = Agenda.objects.filter(archive=archive)
            for ag in agendas:
                ag.is_completed = True
                ag.status = 'selesai'
                ag.completed_notes = f"Laporan Hasil Penanganan ({report.report_number}): {report.title}\n{report.content}"
                if report.file:
                    ag.completed_file = report.file
                ag.save()

        messages.success(request, f"Laporan {report.report_number} berhasil diperbarui.")
        return redirect('reports:detail', pk=report.pk)
    
    return render(request, 'reports/create.html', {
        'dispo': report.disposition,
        'report': report,
        'default_report_number': report.report_number or '',
    })

@login_required
@sdm_required
def drive_backup(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    archive = get_object_or_404(Archive, pk=pk)
    if archive.drive_backed_up:
        messages.info(request, f"Arsip {archive.archive_number} sudah pernah dibackup.")
        return redirect('reports:index')
    result = GoogleIntegrationService.sync_to_drive(archive)
    if result:
        AuditService.log_action(request.user, f"Backup ke Google Drive: {archive.archive_number}", request)
        messages.success(request, f"Arsip {archive.archive_number} berhasil dibackup ke Google Drive.")
    else:
        messages.error(request, f"Gagal backup {archive.archive_number} ke Google Drive. Periksa konfigurasi Drive.")
    return redirect('reports:index')

@login_required
@sdm_required
def backup_test_connection(request):
    results = []
    from services.integrations.google_drive import GoogleDriveService
    svc = GoogleDriveService()

    from reports.models import GoogleOAuthToken
    oauth_token = GoogleOAuthToken.objects.first()
    if oauth_token and oauth_token.refresh_token:
        results.append(("OAuth Login", "Sudah login - refresh token tersimpan"))
    else:
        oauth_config = GoogleOAuthToken.get_client_config()
        if oauth_config:
            results.append(("OAuth Login", "File client ID sudah diupload, tapi belum login"))
        else:
            results.append(("OAuth Login", "Belum dikonfigurasi"))

    results.append(("Service Account File", os.path.exists(svc.creds_path) if svc.creds_path else "Tidak dikonfigurasi"))
    results.append(("Folder ID", svc.folder_id or "Tidak diset"))

    creds = svc._get_credentials()
    if not creds:
        results.append(("Auth Credentials", "Tidak ada credentials valid"))
    else:
        method = "OAuth 2.0" if (oauth_token and oauth_token.refresh_token) else "Service Account"
        results.append(("Auth Method", method))

        from googleapiclient.discovery import build
        drive = build('drive', 'v3', credentials=creds)
        sheets = build('sheets', 'v4', credentials=creds)

        try:
            about = drive.about().get(fields='storageQuota,user').execute()
            quota = about.get('storageQuota', {})
            used = int(quota.get('usage', 0))
            limit = int(quota.get('limit', 0))
            pct = used * 100 // limit if limit else 0
            results.append(("Drive API (storage)", f"OK - {used//1048576}MB / {limit//1048576}MB ({pct}%)"))
        except Exception as e:
            results.append(("Drive API (storage)", f"Gagal: {e}"))

        test_id = None
        try:
            file_meta = {
                'name': '_test_baznas_delete_me',
                'mimeType': 'application/vnd.google-apps.spreadsheet',
            }
            created = drive.files().create(body=file_meta, fields='id').execute()
            test_id = created.get('id')
            results.append(("Buat Spreadsheet via DRIVE API", f"OK - ID: {test_id}"))
        except Exception as e:
            results.append(("Buat Spreadsheet via DRIVE API", f"Gagal: {e}"))

        if test_id:
            try:
                sheets.spreadsheets().values().update(
                    spreadsheetId=test_id,
                    range='Sheet1!A1',
                    valueInputOption='USER_ENTERED',
                    body={'values': [['No', 'Nama'], ['1', 'Test']]}
                ).execute()
                results.append(("Tulis Data via SHEETS API", "OK - Berhasil menulis data"))
            except Exception as e:
                results.append(("Tulis Data via SHEETS API", f"Gagal: {e}"))

            try:
                drive.files().delete(fileId=test_id).execute()
                results.append(("Hapus File Test", "OK"))
            except Exception as e:
                results.append(("Hapus File Test", f"Gagal: {e}"))

        svc._save_token_after_request(creds)

    return render(request, 'reports/diagnostic.html', {'results': results})

@login_required
def backup_count_documents(request):
    month = request.GET.get('month')
    year = request.GET.get('year')
    if not month or not year:
        return JsonResponse({'count': 0, 'error': None})
    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        return JsonResponse({'count': 0, 'error': 'Format tidak valid'})
    if month < 1 or month > 12:
        return JsonResponse({'count': 0, 'error': 'Bulan tidak valid'})

    count = Archive.objects.filter(
        letter_date__month=month,
        letter_date__year=year,
    ).distinct().count()

    return JsonResponse({'count': count, 'error': None})

@login_required
def drive_backup_monthly(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    month = request.POST.get('month')
    year = request.POST.get('year')
    if not month or not year:
        messages.error(request, "Periode bulan dan tahun harus diisi.")
        return redirect('reports:index')

    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        messages.error(request, "Format bulan/tahun tidak valid.")
        return redirect('reports:index')

    if month < 1 or month > 12:
        messages.error(request, "Bulan harus antara 1-12.")
        return redirect('reports:index')

    archives = Archive.objects.filter(
        Q(letter_date__month=month, letter_date__year=year) |
        Q(created_at__month=month, created_at__year=year)
    ).select_related('category', 'uploaded_by').prefetch_related(
        Prefetch('dispositions', queryset=Disposition.objects.select_related('report', 'sender').prefetch_related('forwarded_to', 'waka_forwarded_to', 'sppd_list')),
        'agendas'
    ).distinct().order_by('-created_at')

    if not archives.exists():
        messages.warning(request, f"Tidak ada data laporan untuk bulan {month}/{year}.")
        return redirect('reports:index')

    from services.integrations.google_drive import GoogleDriveService
    drive_service = GoogleDriveService()
    archive_list = list(archives)
    spreadsheet_id, spreadsheet_url, error_msg = drive_service.create_monthly_backup(month, year, archive_list)

    if spreadsheet_id:
        MonthlyBackup.objects.update_or_create(
            month=month, year=year,
            defaults={
                'spreadsheet_id': spreadsheet_id,
                'spreadsheet_url': spreadsheet_url,
            }
        )
        AuditService.log_action(
            request.user,
            f"Backup bulanan {month}/{year} ke Google Sheets ({len(archive_list)} dokumen)",
            request
        )
        messages.success(request, f"Backup bulan {month}/{year} berhasil. {len(archive_list)} dokumen disimpan.")
    else:
        messages.error(request, f"Gagal backup ke Google Sheets. {error_msg or 'Periksa konfigurasi Drive.'}")
    return redirect('reports:index')

@login_required
@sdm_required
def oauth_login(request):
    import os, json, secrets, hashlib, base64
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    from reports.models import GoogleOAuthToken

    client_config = GoogleOAuthToken.get_client_config()
    if not client_config:
        messages.error(request, "Upload OAuth Client ID JSON dulu di Pengaturan Aplikasi.")
        return redirect('reports:index')

    cfg = client_config.get('installed', client_config.get('web', client_config))
    client_id = cfg.get('client_id')
    if not client_id:
        messages.error(request, "client_id tidak ditemukan di file JSON.")
        return redirect('reports:index')

    redirect_uris = cfg.get('redirect_uris', [])
    current_redirect = request.build_absolute_uri('/reports/oauth/callback/')
    
    # Cocokkan redirect_uri secara dinamis sesuai host browser pengguna (127.0.0.1 vs localhost)
    matching_uri = None
    for u in redirect_uris:
        if u.rstrip('/') == current_redirect.rstrip('/'):
            matching_uri = u
            break
    
    if matching_uri:
        redirect_uri = matching_uri
    elif '127.0.0.1' in request.get_host() and any('127.0.0.1' in u for u in redirect_uris):
        redirect_uri = [u for u in redirect_uris if '127.0.0.1' in u][0]
    elif 'localhost' in request.get_host() and any('localhost' in u for u in redirect_uris):
        redirect_uri = [u for u in redirect_uris if 'localhost' in u][0]
    else:
        redirect_uri = current_redirect

    scopes = ['https://www.googleapis.com/auth/drive.file',
              'https://www.googleapis.com/auth/spreadsheets']

    code_verifier = secrets.token_urlsafe(32)
    digest = hashlib.sha256(code_verifier.encode('ascii')).digest()
    code_challenge = base64.urlsafe_b64encode(digest).rstrip(b'=').decode('ascii')

    from requests_oauthlib import OAuth2Session
    session = OAuth2Session(client_id, redirect_uri=redirect_uri, scope=scopes)
    auth_url, state = session.authorization_url(
        'https://accounts.google.com/o/oauth2/auth',
        code_challenge=code_challenge,
        code_challenge_method='S256',
        access_type='offline',
        prompt='consent'
    )
    request.session['oauth_code_verifier'] = code_verifier
    request.session['oauth_state'] = state
    request.session['oauth_redirect_uri'] = redirect_uri
    return redirect(auth_url)

@login_required
@sdm_required
def oauth_callback(request):
    import os, json
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    from reports.models import GoogleOAuthToken

    code_verifier = request.session.get('oauth_code_verifier')
    if not code_verifier:
        messages.error(request, "Sesi OAuth tidak ditemukan. Silakan klik 'Login dengan Google' kembali.")
        return redirect('users:app_settings')

    client_config = GoogleOAuthToken.get_client_config()
    if not client_config:
        messages.error(request, "Konfigurasi OAuth tidak ditemukan.")
        return redirect('users:app_settings')

    cfg = client_config.get('installed', client_config.get('web', client_config))
    client_id = cfg.get('client_id')
    client_secret = cfg.get('client_secret')
    if not client_id or not client_secret:
        messages.error(request, "client_id atau client_secret tidak ditemukan.")
        return redirect('users:app_settings')

    redirect_uri = request.session.get('oauth_redirect_uri') or request.build_absolute_uri('/reports/oauth/callback/')
    scopes = ['https://www.googleapis.com/auth/drive.file',
              'https://www.googleapis.com/auth/spreadsheets']

    from requests_oauthlib import OAuth2Session
    session = OAuth2Session(client_id, redirect_uri=redirect_uri, scope=scopes)
    token = session.fetch_token(
        'https://oauth2.googleapis.com/token',
        client_secret=client_secret,
        authorization_response=request.build_absolute_uri(),
        code_verifier=code_verifier
    )

    token_obj, _ = GoogleOAuthToken.objects.get_or_create(pk=1)
    token_obj.refresh_token = token.get('refresh_token', '') or token_obj.refresh_token
    token_obj.access_token = token.get('access_token', '')
    from django.utils import timezone
    from datetime import timedelta
    expires_in = token.get('expires_in', 3600)
    token_obj.token_expiry = timezone.now() + timedelta(seconds=expires_in)
    token_obj.save()

    if 'oauth_code_verifier' in request.session:
        del request.session['oauth_code_verifier']
    if 'oauth_state' in request.session:
        del request.session['oauth_state']
    messages.success(request, "Berhasil terhubung ke Google Drive BAZNAS! Fitur pencadangan otomatis telah aktif 100%.")
    return redirect('users:app_settings')

@login_required
@sdm_required
def drive_backup_batch(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    ids = request.POST.getlist('ids')
    if not ids:
        messages.warning(request, "Tidak ada arsip yang dipilih.")
        return redirect('reports:index')
    success_count = 0
    fail_count = 0
    archives = Archive.objects.filter(pk__in=ids, drive_backed_up=False)
    for archive in archives:
        result = GoogleIntegrationService.sync_to_drive(archive)
        if result:
            AuditService.log_action(request.user, f"Backup batch ke Google Drive: {archive.archive_number}", request)
            success_count += 1
        else:
            fail_count += 1
    if success_count:
        messages.success(request, f"{success_count} arsip berhasil dibackup ke Google Drive.")
    if fail_count:
        messages.warning(request, f"{fail_count} arsip gagal dibackup.")
    return redirect('reports:index')

@login_required
def rekap_sppd_view(request):
    from services.analytics.reporting_service import ReportingService
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    sppd_recap = ReportingService.get_sppd_recap(year=year, month=month)
    return render(request, 'reports/rekap_sppd.html', {
        'sppd_recap': sppd_recap,
        'year': year,
        'month': month,
        'month_names': ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'],
    })

@login_required
def rekap_bantuan_view(request):
    from services.analytics.reporting_service import ReportingService
    import json
    year_param = request.GET.get('year')
    month_param = request.GET.get('month')

    if not year_param and not month_param:
        all_time = True
        year = None
        month = None
    elif year_param == 'all' or month_param == 'all':
        all_time = True
        year = None
        month = None
    else:
        all_time = False
        year = int(year_param) if year_param and year_param.isdigit() else None
        month = int(month_param) if month_param and month_param.isdigit() else None

    active_pov = request.session.get('active_pov')
    is_waka_or_kabid_2 = active_pov in ['waka_2', 'kabid_2'] or getattr(request.user, 'is_waka_2', False) or getattr(request.user, 'is_kabid_2', False)
    if not is_waka_or_kabid_2:
        emp = getattr(request.user, 'employee', None)
        if emp and emp.dept_relation and ('pendistribusian' in emp.dept_relation.name.lower() or 'bidang ii' in emp.dept_relation.name.lower() or 'bidang 2' in emp.dept_relation.name.lower()):
            is_waka_or_kabid_2 = True

    bantuan_analytics = ReportingService.get_bantuan_analytics(year=year, month=month, all_time=all_time)
    
    return render(request, 'reports/rekap_bantuan.html', {
        'analytics': bantuan_analytics,
        'bantuan_chart_labels': bantuan_analytics['bantuan_chart_labels'],
        'bantuan_chart_series': bantuan_analytics['bantuan_chart_series'],
        'umum_chart_labels': bantuan_analytics['umum_chart_labels'],
        'umum_chart_series': bantuan_analytics['umum_chart_series'],
        'selected_year': year or timezone.now().year,
        'selected_month': month or timezone.now().month,
        'all_time': all_time,
        'is_waka_or_kabid_2': is_waka_or_kabid_2,
        'month_names': [
            (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
            (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
            (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
        ],
    })

@login_required
def export_rekap_bantuan_excel(request):
    """
    Ekspor Laporan Rekapitulasi Penanganan Dokumen Bantuan & Dokumen Umum ke Excel (.xlsx).
    Dilengkapi Kop Surat Resmi Logo BAZNAS di sebelah kiri, All Border, & Wrap Text.
    """
    from services.analytics.reporting_service import ReportingService
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    year_param = request.GET.get('year')
    month_param = request.GET.get('month')
    all_time_param = request.GET.get('all_time')

    if all_time_param == '1' or (not year_param and not month_param) or year_param == 'all' or month_param == 'all':
        all_time = True
        year = None
        month = None
        periode_str = "Semua_Data_Histori"
    else:
        all_time = False
        year = int(year_param) if year_param and year_param.isdigit() else timezone.now().year
        month = int(month_param) if month_param and month_param.isdigit() else timezone.now().month
        month_names_dict = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
            5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
            9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        m_name = month_names_dict.get(month, str(month))
        periode_str = f"{m_name}_{year}"

    analytics = ReportingService.get_bantuan_analytics(year=year, month=month, all_time=all_time)

    wb = Workbook()
    logo_path = _get_logo_path()

    font_kop_title = Font(name='Calibri', size=13, bold=True, color='006633')
    font_kop_sub = Font(name='Calibri', size=11, bold=True, color='1E293B')
    font_kop_meta = Font(name='Calibri', size=9.5, italic=True, color='475569')

    fill_header_green = PatternFill(start_color='006633', end_color='006633', fill_type='solid')
    fill_header_blue = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    font_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    fill_subhead = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    font_subhead = Font(name='Calibri', size=10, bold=True, color='0F172A')

    thin_gray = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def apply_kop_surat(ws, sheet_title_text):
        ws.views.sheetView[0].showGridLines = True
        
        ws.row_dimensions[1].height = 6
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 6

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 16
        ws.merge_cells('B2:B4')
        ws['B2'].alignment = align_center_wrap

        if logo_path:
            try:
                img = OpenPyXLImage(logo_path)
                img.height = 54
                img.width = 88
                ws.add_image(img, 'B2')
            except Exception:
                pass

        ws.merge_cells('C2:F2')
        ws['C2'].value = "BADAN AMIL ZAKAT NASIONAL (BAZNAS) KABUPATEN TANGERANG"
        ws['C2'].font = font_kop_title
        ws['C2'].alignment = align_left_wrap

        ws.merge_cells('C3:F3')
        ws['C3'].value = sheet_title_text
        ws['C3'].font = font_kop_sub
        ws['C3'].alignment = align_left_wrap

        ws.merge_cells('C4:F4')
        ws['C4'].value = f"Periode: {periode_str.replace('_', ' ')} | Diekspor: {timezone.now().strftime('%d/%m/%Y %H:%M')} WIB"
        ws['C4'].font = font_kop_meta
        ws['C4'].alignment = align_left_wrap

        for r in range(2, 5):
            for c in range(2, 7):
                ws.cell(row=r, column=c).border = thin_border

        double_bottom = Border(bottom=Side(style='double', color='006633'))
        for c in range(2, 7):
            ws.cell(row=5, column=c).border = double_bottom

    # ---------------------------------------------------------
    # SHEET 1: PERMOHONAN BANTUAN
    # ---------------------------------------------------------
    ws_bantuan = wb.active
    ws_bantuan.title = "Permohonan Bantuan"
    apply_kop_surat(ws_bantuan, "LAPORAN REKAPITULASI PENANGANAN PERMOHONAN BANTUAN (MUSTAHIK)")

    ws_bantuan.merge_cells('B7:F7')
    ws_bantuan['B7'].value = "RINGKASAN STATISTIK DOKUMEN & BANTUAN"
    ws_bantuan['B7'].font = font_subhead
    ws_bantuan['B7'].fill = fill_subhead

    ws_bantuan.cell(row=8, column=2, value="Total Volume Dokumen Masuk").font = Font(bold=True)
    ws_bantuan.cell(row=8, column=3, value=analytics['total_documents']).alignment = align_center_wrap
    ws_bantuan.cell(row=8, column=5, value="Total Permohonan Bantuan").font = Font(bold=True)
    ws_bantuan.cell(row=8, column=6, value=analytics['total_bantuan']).alignment = align_center_wrap

    ws_bantuan.cell(row=9, column=2, value="Bantuan Selesai Terekap").font = Font(bold=True)
    ws_bantuan.cell(row=9, column=3, value=analytics['bantuan_completed']).alignment = align_center_wrap
    ws_bantuan.cell(row=9, column=5, value="Tingkat Penyelesaian Selesai").font = Font(bold=True)
    ws_bantuan.cell(row=9, column=6, value=f"{analytics['completion_rate']}%").alignment = align_center_wrap

    for r in range(7, 10):
        for c in range(2, 7):
            ws_bantuan.cell(row=r, column=c).border = thin_border

    bantuan_headers = ['NO', 'TANGGAL MASUK', 'NO. DOKUMEN', 'PERIHAL / JUDUL PROPOSAL', 'SUB-KATEGORI BANTUAN', 'STATUS WORKFLOW']
    row_idx = 11
    ws_bantuan.row_dimensions[row_idx].height = 25

    for col_i, h_text in enumerate(bantuan_headers, 1):
        c = ws_bantuan.cell(row=row_idx, column=col_i, value=h_text)
        c.fill = fill_header_green
        c.font = font_header
        c.alignment = align_center_wrap
        c.border = thin_border

    for idx, item in enumerate(analytics['bantuan_details'], 1):
        row_idx += 1
        arc = item['archive']
        ws_bantuan.row_dimensions[row_idx].height = 28

        c1 = ws_bantuan.cell(row=row_idx, column=1, value=idx)
        c1.alignment = align_center_wrap

        c2 = ws_bantuan.cell(row=row_idx, column=2, value=arc.created_at.strftime('%d/%m/%Y'))
        c2.alignment = align_center_wrap

        c3 = ws_bantuan.cell(row=row_idx, column=3, value=arc.archive_number or '(DRAFT)')
        c3.alignment = align_center_wrap

        c4 = ws_bantuan.cell(row=row_idx, column=4, value=arc.title or '-')
        c4.alignment = align_left_wrap

        c5 = ws_bantuan.cell(row=row_idx, column=5, value=item['sub_category'])
        c5.alignment = align_center_wrap

        c6 = ws_bantuan.cell(row=row_idx, column=6, value=arc.workflow_status_display)
        c6.alignment = align_center_wrap

        for col_i in range(1, 7):
            ws_bantuan.cell(row=row_idx, column=col_i).border = thin_border

    ws_bantuan.column_dimensions['A'].width = 6
    ws_bantuan.column_dimensions['B'].width = 16
    ws_bantuan.column_dimensions['C'].width = 22
    ws_bantuan.column_dimensions['D'].width = 42
    ws_bantuan.column_dimensions['E'].width = 26
    ws_bantuan.column_dimensions['F'].width = 28

    # ---------------------------------------------------------
    # SHEET 2: DOKUMEN UMUM
    # ---------------------------------------------------------
    ws_umum = wb.create_sheet(title="Dokumen Umum")
    apply_kop_surat(ws_umum, "LAPORAN REKAPITULASI PENANGANAN SURAT & DOKUMEN UMUM")

    ws_umum.merge_cells('B7:F7')
    ws_umum['B7'].value = "RINGKASAN STATISTIK DOKUMEN UMUM"
    ws_umum['B7'].font = font_subhead
    ws_umum['B7'].fill = fill_subhead

    ws_umum.cell(row=8, column=2, value="Total Dokumen Umum").font = Font(bold=True)
    ws_umum.cell(row=8, column=3, value=analytics['total_umum']).alignment = align_center_wrap
    ws_umum.cell(row=8, column=5, value="Dokumen Umum Selesai").font = Font(bold=True)
    ws_umum.cell(row=8, column=6, value=analytics['umum_completed']).alignment = align_center_wrap

    ws_umum.cell(row=9, column=2, value="Total Dokumen Masuk").font = Font(bold=True)
    ws_umum.cell(row=9, column=3, value=analytics['total_documents']).alignment = align_center_wrap
    ws_umum.cell(row=9, column=5, value="Persentase Penyelesaian").font = Font(bold=True)
    ws_umum.cell(row=9, column=6, value=f"{analytics['completion_rate']}%").alignment = align_center_wrap

    for r in range(7, 10):
        for c in range(2, 7):
            ws_umum.cell(row=r, column=c).border = thin_border

    umum_headers = ['NO', 'TANGGAL MASUK', 'NO. DOKUMEN', 'PERIHAL / SURAT DINAS', 'KLASIFIKASI UMUM', 'STATUS WORKFLOW']
    row_idx = 11
    ws_umum.row_dimensions[row_idx].height = 25

    for col_i, h_text in enumerate(umum_headers, 1):
        c = ws_umum.cell(row=row_idx, column=col_i, value=h_text)
        c.fill = fill_header_blue
        c.font = font_header
        c.alignment = align_center_wrap
        c.border = thin_border

    for idx, item in enumerate(analytics['umum_details'], 1):
        row_idx += 1
        arc = item['archive']
        ws_umum.row_dimensions[row_idx].height = 28

        c1 = ws_umum.cell(row=row_idx, column=1, value=idx)
        c1.alignment = align_center_wrap

        c2 = ws_umum.cell(row=row_idx, column=2, value=arc.created_at.strftime('%d/%m/%Y'))
        c2.alignment = align_center_wrap

        c3 = ws_umum.cell(row=row_idx, column=3, value=arc.archive_number or '(DRAFT)')
        c3.alignment = align_center_wrap

        c4 = ws_umum.cell(row=row_idx, column=4, value=arc.title or '-')
        c4.alignment = align_left_wrap

        c5 = ws_umum.cell(row=row_idx, column=5, value=item['sub_category'])
        c5.alignment = align_center_wrap

        c6 = ws_umum.cell(row=row_idx, column=6, value=arc.workflow_status_display)
        c6.alignment = align_center_wrap

        for col_i in range(1, 7):
            ws_umum.cell(row=row_idx, column=col_i).border = thin_border

    ws_umum.column_dimensions['A'].width = 6
    ws_umum.column_dimensions['B'].width = 16
    ws_umum.column_dimensions['C'].width = 22
    ws_umum.column_dimensions['D'].width = 42
    ws_umum.column_dimensions['E'].width = 26
    ws_umum.column_dimensions['F'].width = 28

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename_str = f"SIMAP_Rekap_Penanganan_BAZNAS_{periode_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'
    wb.save(response)
    return response

@login_required
def calendar_work_view(request):
    from services.calendar.calendar_service import CalendarService
    events = CalendarService.get_calendar_events()
    import json
    return render(request, 'calendar/index.html', {
        'events': events,
        'events_json': json.dumps(events),
    })

# Alias for compatibility
reports_index = report_index